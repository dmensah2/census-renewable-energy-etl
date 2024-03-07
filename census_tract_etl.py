#Deidre Mensah

#import modules
from arcgis.gis import GIS
from pathlib import Path
import geopandas as gpd
import py7zr as py7
from sqlalchemy import create_engine
import psycopg2
from psycopg2 import sql
import requests
from zipfile import ZipFile
from openpyxl import load_workbook

#CENSUS TRACT EXTRACT
def census_tract_extract(publicid, folder_name, lpk_file, gdb):
    public_id = publicid

    gis = GIS()

    #https request to retrieve arcgis item
    data_item = gis.content.get(public_id)

    #create path object to create a data folder in the current path the project is located in
    data_path = Path('./' + folder_name)

    #check to see if data folder doesn't exist yet
    if not data_path.exists():
        #if data folder doesn't exist, create new folder
        data_path.mkdir()

    #download item
    data_item.download(save_path=data_path)

    #create a file path to the object
    zip_path = data_path.joinpath(lpk_file)

    print ("extracting spatial features....")
    #extract spatial features from the object in the file path
    with py7.SevenZipFile(zip_path, mode='r') as z:
        z.extractall()

    print ("extracting completed")
    #create a new file path
    gdb_file_path = Path('./v107/' + gdb)

    #read the .gdb file in the file path
    gdf = gpd.read_file(gdb_file_path)

    return gdf

#EIA 860 EXTRACT
def eia_860_extract(url, folder_path, save_path, chunk_size=128):
    # create path object to create a data folder in the current path the project is located in
    data_path = Path('./' + folder_path)

    # check to see if data folder doesn't exist yet
    if not data_path.exists():
        # if data folder doesn't exist, create new folder
        data_path.mkdir()

    #get url and store content in memory instead of downloading
    response = requests.get(url, stream=True)

    #save response content progressively in chunks
    with open("./" + folder_path + "/" + save_path, "wb") as fd:
        for chunk in response.iter_content(chunk_size=chunk_size):
            fd.write(chunk)

    print("extracting spreadsheets...")
    #checks to make sure file is a .zip file
    if save_path.lower().endswith('.zip'):
        #extract files
        with ZipFile("./" + folder_path + "/" + save_path, "r") as z:
            #extract to local data folder
            z.extractall(path='./' + folder_path)

    print("extracting completed")
    return


#LOAD DATA
def load_data_to_postgis(gdf, source_folder, new_db_name, postgres_db, postgres_user, postgres_pw, postgres_host, postgres_port):

    #CONNECT TO POSTGRES
    print("...connecting to DB")
    #connect to default postgres db
    conn = psycopg2.connect(
        dbname=postgres_db,
        user=postgres_user,
        password=postgres_pw,
        host=postgres_host,
        port=str(postgres_port)
    )

    # Set autocommit to True
    conn.autocommit = True

    #execute queries in default db
    with conn.cursor() as cursor:
        #enable postGIS extension
        cursor.execute("CREATE EXTENSION IF NOT EXISTS postgis")

        #define the engine
        engine = create_engine("postgresql://" + postgres_user + ":" + postgres_pw + "@" + postgres_host + ":" + str(postgres_port) + "/" + new_db_name)

        #create query to check if database currently exists=
        query = "SELECT datname FROM pg_database WHERE lower(datname) = lower('" + new_db_name + "');"
        cursor.execute(query, (new_db_name,))
        result = cursor.fetchone()

        #if database exists add new table
        if result:
            print(f"Database '{new_db_name}' already exists.")
            print(f"loading census tracts...")
            gdf.to_postgis("us_census_tracts", engine, if_exists='replace', index=False)

        #create new table first then add the table
        else:
            print(f"Database '{new_db_name}' does not exist.")

            print(f"Creating new database...")
            # create the new database
            cursor.execute(sql.SQL("CREATE DATABASE {}").format(sql.Identifier(new_db_name)))

    # Set autocommit back to False
    conn.autocommit = False

    # Close the connection
    conn.close()

    # connect to new postgres db
    new_conn = psycopg2.connect(
        dbname=new_db_name,
        user=postgres_user,
        password=postgres_pw,
        host=postgres_host,
        port=str(postgres_port)
    )

    # Set autocommit to True
    new_conn.autocommit = True

    ##LOAD CENSUS TRACTS
    #create postgis extension in new dB
    with new_conn.cursor() as new_cursor:

        # define the engine
        engine = create_engine("postgresql://" + postgres_user + ":" + postgres_pw + "@" + postgres_host + ":" + str(postgres_port) + "/" + new_db_name)

        #check to see if postgis is enabled
        new_cursor.execute("SELECT 1 FROM pg_extension WHERE extname='postgis';")
        pg_result = new_cursor.fetchone()

        #if postgis extension does not exist
        if pg_result is None:
            print("...connecting to new database")
            # create new postgis extension in database
            new_cursor.execute(sql.SQL("CREATE EXTENSION postgis;").format(sql.Identifier(new_db_name)))
            #load census tracts
            print(f"loading census tracts...")
            gdf.to_postgis("us_census_tracts", engine, if_exists='replace', index=False)

        ##LOAD SPREADSHEETS
        plant = load_workbook("./" + source_folder + "/2___Plant_Y2022.xlsx")

        #get specific sheet by name
        generator = load_workbook("./" + source_folder + "/3_1_Generator_Y2022.xlsx")
        sheet_name = "Retired and Canceled"
        generator_sheet = generator[sheet_name]

        plant_sheet = plant.active
        #generator_sheet = generator_sheet.active

        # Create a list with the column names in the first row of the workbook
        plant_column_names = [column.value for column in plant_sheet[2]]
        generator_column_names = [column.value for column in generator_sheet[2]]

        plant_data = []
        generator_data = []

        #iterate through rows in sheet to extract data and append to list
        for row in plant_sheet.iter_rows(min_row=3, values_only=True):
            plant_data.append(row)
        for row in generator_sheet.iter_rows(min_row=3, values_only=True):
            generator_data.append(row)

        # Set a name for the PostgreSQL schema and table where we will put the data
        schema_name = 'energy'
        plant_table_name = 'Plant_Y2022'
        generator_table_name = 'Generator_Y2022'

        # Write a query to create a schema using schema_name
        schema_creation_query = f'CREATE SCHEMA IF NOT EXISTS {schema_name}'

        # Write a query to create a table in the schema. It must contain all
        # columns in column_names
        plant_table_creation_query = f"""
         CREATE TABLE IF NOT EXISTS {schema_name}.{plant_table_name} (
         {", ".join([f'"{name}" TEXT' for name in plant_column_names])}
         )
        """

        generator_table_creation_query = f"""
         CREATE TABLE IF NOT EXISTS {schema_name}.{generator_table_name} (
         {", ".join([f'"{name}" TEXT' for name in generator_column_names])}
         )
        """

        print("loading spreadsheets...")
        # Insert data into PostgreSQL table
        new_cursor.execute(schema_creation_query)
        new_cursor.execute(plant_table_creation_query)
        new_cursor.execute(generator_table_creation_query)

        # Create a parameterized SQL query to insert the data into the table
        insert_plant_data_query = f"""
           INSERT INTO {schema_name}.{plant_table_name} ({", ".join([f'"{name}"' for name in plant_column_names])})
           VALUES ({", ".join(['%s' for _ in plant_column_names])})
        """

        # Create a parameterized SQL query to insert the data into the table
        insert_generator_data_query = f"""
          INSERT INTO {schema_name}.{generator_table_name} ({", ".join([f'"{name}"' for name in generator_column_names])})
          VALUES ({", ".join(['%s' for _ in generator_column_names])})
        """

        # Execute the query using the data list as parameter
        new_cursor.executemany(insert_plant_data_query, plant_data)
        new_cursor.executemany(insert_generator_data_query, generator_data)

        # Set autocommit back to False
        new_conn.autocommit = False

        # Close the connection
        new_conn.close()

        # Print a message
        print('Import successfully completed!')
